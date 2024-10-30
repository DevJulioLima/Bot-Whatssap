import openpyxl
from urllib.parse import quote

clientes = openpyxl.load_workbook("clientes.xlsx")
pagina_clientes = clientes['Planilha1']

produtos = openpyxl.load_workbook("produtos.xlsx")
pagina_produtos = produtos['Planilha1']

for linha in pagina_clientes.iter_rows(min_row=2):
    # nome e telefone.
    nome = linha[0].value
    telefone = linha[1].value
    
for linha in pagina_produtos.iter_rows(min_row=2):
    # produto, preço e descrição
    produto = linha[0].value
    preco = linha[1].value
    descricao = linha[2].value
    mensagem = f'ola {nome} esses são os produtos disponoiveis {produto} preço {preco} e descrição: {descricao}'
    link_mensagem_whatssap = f'https://web.whatssap.com/send?phone={telefone}&text={quote(mensagem)}'
